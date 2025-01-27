import openpyxl

def get_row_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_row

def get_column_count(file, sheet):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.max_column

def read_data(file, sheet, row_num, column_num):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    return sheet.cell(row_num, column_num).value

def write_data(file, sheet, row_num, column_num, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheet]
    sheet.cell(row_num, column_num).value = data
    workbook.save(file)

